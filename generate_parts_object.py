import openpyxl


def load_workbook_data():

    def evaluate_characteristics(identity,string):
        assert type(identity) is list, "There was an issue evaluating the excel sheet. Check identities parameter"

        if len(string) < 3:
            return ["Georgia","Unknown","Unknown","Unknown"]
        return_list = [[]]*len(identity)
        # Validate that the characteristics are of the correct format. Return default expression if false
        for index,i in enumerate(string):
            if not i.isalpha():
                break
            string_length = index + 1
        try:
            string_length += 0
        except:
            string_length = 1

        if string_length == 5:
            if string[0] in identity[0]:
                return_list[0] = identities[0][string[0]]
            else:
                return_list[0] = "Unknown"
            if string[1:3]  in identity[1]:
                return_list[1] = identities[1][string[1:3]]
            else:
                return_list[1] = "Unknown"
            if string[3] in identity[2]:
                return_list[2] = identities[2][string[3]]
            else:
                return_list[2] = "Unknown"
            if string[4] in identity[3]:
                return_list[3] = identities[3][string[4]]
            else:
                return_list[3] = "Unknown"
            return return_list

        elif string_length == 6:
            if string[0] in identity[0]:
                return_list[0] = identities[0][string[0]]
            else:
                return_list[0] = "Unknown"
            if string[1:4] in identity[1]:
                return_list[1] = identities[1][string[1:4]]
            else:
                return_list[1] = "Unknown"
            if string[4] in identity[2]:
                return_list[2] = identities[2][string[4]]
            else:
                return_list[2] = "Unknown"
            if string[5] in identity[3]:
                return_list[3] = identities[3][string[5]]
            else:
                return_list[3] = "Unknown"
            return return_list
        return ["Georgia", "Unknown", "Unknown", "Unknown"]

    # Location, machine, type, use
    identities = [{}]*4
    identities[0] = {"A":"JC","B":"ES","C":"Georgia"}
    identities[1] = {
        "CT": "Cleaving",
        "CA": "Cleaving",
        "TB": "Tabber",
        "LU": "Lay Up",
        "FD": "Front Conveyor",
        "AS": "Auto Solder",
        "GR": "Glass Loader",
        "FEI": "Front EL",
        "MA": "Matrix",
        "SC": "Sheet Cutting",
        "LM": "Laminator",
        "MD": "Conveyor",
        "TR": "Trimming",
        "VI": "Vision Inspection",
        "JD": "Silicone Dispenser",
        "JB": "JB Solder",
        "FA": "Frame Assembly",
        "FP": "Frame Press",
        "MFA": "Middle Frane Assembly",
        "FG": "Frame Grinder",
        "SM": "Simulator",
        "HV": "HV Tester",
        "EEI": "Back EL",
        "PO": "Potting",
        "CU": "Curing",
        "GT": "Ground Tester",
        "PL": "Power Label",
        "SL": "Side Label",
        "FC": "FQC Cleaning",
        "CP": "Corner Protector",
        "GD": "Grading",
        "PA": "Packing",
        "LG": "LGV",
        "ED": "Back Conveyor",
        "UT": "Utility",
        "CM": "Common",
        "__": "Unknown"
    }
    identities[2] = {"E":"Electric", "A": "Pneumatic", "M":"Machinery","C":"Common"}
    identities[3] = {
        "C": "Consumable",
        "S": "Spare"
    }


    ##
    wb = openpyxl.load_workbook(filename='RawPartslistXL/list of parts.XLSX')
    ws = wb.active
    workbook_pre_database = {}
    # loop through the worksheet collect Material code, Material description, basic material code -> [identities]
    #
    r, c = 2, 1
    # Items on the parts list excel file start at row 2
    while ws.cell(row=r, column=1).value is not None:
        # Generate objects
        characteristics = evaluate_characteristics(identities,ws.cell(row=r, column=17).value)
        workbook_pre_database[f"{ws.cell(row=r, column=1).value}"] = {
            "Description":ws.cell(row=r, column=2).value,
            "Creator":ws.cell(row=r, column=4).value,
            "Machine":characteristics[1],
            "Location":characteristics[0],
            "type":characteristics[2],
            "usage":characteristics[3],
            "Company": ws.cell(row=r, column=16).value if ws.cell(row=r, column=16).value != "" else "Unknown",
            "Vendor ID":ws.cell(row=r, column=18).value if ws.cell(row=r, column=18).value != "" else "Unknown"
        }
        # Location, machine, type, use
        r += 1
    return workbook_pre_database

